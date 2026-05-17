import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import requests
import json
import io
import re

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="OMAC AI Complaint Intelligence Tool",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Constants ─────────────────────────────────────────────────────────────────
GROQ_API_KEY = "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"
GROQ_URL     = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL   = "llama3-70b-8192"
SAFE_BATCH   = 3

# ── Session defaults ──────────────────────────────────────────────────────────
for k, v in {
    "wb_bytes": None, "file_name": "", "sheet_name": "",
    "df": None, "result_df": None, "log": [],
    "dark": True, "use_api": True,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Theme vars ────────────────────────────────────────────────────────────────
D = st.session_state.dark
BG      = "#0d1117"   if D else "#f4f6fb"
SURFACE = "#161b22"   if D else "#ffffff"
SURFACE2= "#1c2333"   if D else "#f0f4ff"
BORDER  = "#30363d"   if D else "#d0d7de"
TEXT    = "#e6edf3"   if D else "#1a1f36"
TEXT2   = "#8b949e"   if D else "#6e7891"
ACCENT  = "#3b82f6"
ACCENT2 = "#6366f1"
GREEN   = "#22c55e"   if D else "#16a34a"
RED     = "#ef4444"   if D else "#dc2626"
YELLOW  = "#f59e0b"   if D else "#d97706"

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] {{
    font-family: 'Inter', sans-serif;
    background-color: {BG};
    color: {TEXT};
}}

.stApp {{ background-color: {BG}; }}

/* ── Header ── */
.omac-header {{
    background: linear-gradient(135deg, #1e3a8a 0%, #3b5bdb 60%, {ACCENT2} 100%);
    border-radius: 16px;
    padding: 26px 32px;
    margin-bottom: 28px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 4px 24px rgba(59,130,246,0.25);
}}
.omac-title {{ color:#fff; font-size:24px; font-weight:800; margin:0; letter-spacing:-0.3px; }}
.omac-sub   {{ color:#bfdbfe; font-size:13px; margin:4px 0 0; font-weight:400; }}
.omac-badge {{
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.25);
    border-radius: 8px;
    padding: 8px 14px;
    color: #e0eaff;
    font-size: 12px;
    text-align: right;
    backdrop-filter: blur(4px);
}}

/* ── Section headers ── */
.section-head {{
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 28px 0 14px;
}}
.step-num {{
    background: linear-gradient(135deg, {ACCENT}, {ACCENT2});
    color: #fff;
    border-radius: 50%;
    width: 28px; height: 28px;
    line-height: 28px;
    text-align: center;
    font-weight: 700;
    font-size: 13px;
    flex-shrink: 0;
    box-shadow: 0 2px 8px rgba(99,102,241,0.4);
}}
.step-label {{ font-weight: 700; font-size: 15px; color: {TEXT}; }}

/* ── Cards ── */
.card {{
    background: {SURFACE};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 16px 20px;
    margin-bottom: 10px;
}}
.info-card  {{ background: {'#0d2137' if D else '#eff6ff'}; border-color: {'#1d4ed8' if D else '#bfdbfe'}; color: {'#93c5fd' if D else '#1e40af'}; border-radius:10px; padding:12px 16px; font-size:13px; margin-bottom:8px; }}
.success-card {{ background: {'#052e16' if D else '#f0fdf4'}; border: 1px solid {'#166534' if D else '#86efac'}; border-radius:10px; padding:14px 18px; color: {GREEN}; font-size:14px; }}
.warn-card  {{ background: {'#1c1300' if D else '#fffbeb'}; border: 1px solid {'#92400e' if D else '#fcd34d'}; border-radius:10px; padding:12px 16px; font-size:13px; color: {YELLOW}; margin-bottom:8px; }}

/* ── Severity pills ── */
.pill {{
    display: inline-block;
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 12px;
    font-weight: 600;
    margin: 2px;
}}
.pill-serious  {{ background:#fee2e2; color:#991b1b; }}
.pill-major    {{ background:#ffedd5; color:#9a3412; }}
.pill-moderate {{ background:#fef9c3; color:#854d0e; }}
.pill-minor    {{ background:#dcfce7; color:#166534; }}
.pill-harm-yes {{ background:#fecaca; color:#7f1d1d; }}
.pill-harm-no  {{ background:#d1fae5; color:#065f46; }}

/* ── Mode toggle strip ── */
.mode-strip {{
    display: flex;
    gap: 8px;
    margin-bottom: 16px;
}}
.mode-btn {{
    flex: 1;
    text-align: center;
    padding: 10px;
    border-radius: 10px;
    font-size: 13px;
    font-weight: 600;
    cursor: pointer;
    border: 2px solid {BORDER};
    background: {SURFACE};
    color: {TEXT2};
}}
.mode-btn.active {{
    border-color: {ACCENT};
    background: {'#0d2137' if D else '#eff6ff'};
    color: {ACCENT};
}}

/* ── Log box ── */
.log-box {{
    background: #0d1117;
    border: 1px solid #30363d;
    border-radius: 10px;
    padding: 14px 16px;
    font-family: 'JetBrains Mono', 'Fira Code', monospace;
    font-size: 12px;
    max-height: 280px;
    overflow-y: auto;
    margin-top: 12px;
}}
.log-success {{ color: #4ade80; }}
.log-error   {{ color: #f87171; }}
.log-info    {{ color: #94a3b8; }}

/* ── Footer ── */
.omac-footer {{
    text-align: center;
    color: {TEXT2};
    font-size: 12px;
    margin-top: 48px;
    padding-top: 18px;
    border-top: 1px solid {BORDER};
}}
.omac-footer strong {{ color: {ACCENT}; }}

/* ── Streamlit overrides ── */
div[data-testid="stSelectbox"] > div,
div[data-testid="stMultiSelect"] > div {{
    background: {SURFACE2} !important;
    border-color: {BORDER} !important;
    color: {TEXT} !important;
    border-radius: 8px !important;
}}
div[data-testid="stTextArea"] textarea {{
    background: {SURFACE2} !important;
    border-color: {BORDER} !important;
    color: {TEXT} !important;
    border-radius: 8px !important;
}}
div[data-testid="stMetric"] {{
    background: {SURFACE};
    border: 1px solid {BORDER};
    border-radius: 10px;
    padding: 12px 16px;
}}
hr {{ border-color: {BORDER}; }}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# HEADER + TOGGLE ROW
# ═══════════════════════════════════════════════════════════════════════════════
hcol1, hcol2 = st.columns([4, 1])
with hcol1:
    st.markdown("""
    <div class="omac-header">
      <div>
        <div class="omac-title">🏥 OMAC Complaint Intelligence</div>
        <div class="omac-sub">AI-powered severity · affinity · harm tagging for complaint registers</div>
      </div>
      <div class="omac-badge">
        Powered by Groq LLaMA 3<br>
        <span style="color:#93c5fd; font-weight:600;">llama3-70b-8192</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

with hcol2:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    if st.button("🌙 Dark" if not D else "☀️ Light", use_container_width=True):
        st.session_state.dark = not st.session_state.dark
        st.rerun()
    api_label = "⚡ API Mode ON" if st.session_state.use_api else "🔧 Rule-Based Mode"
    if st.button(api_label, use_container_width=True):
        st.session_state.use_api = not st.session_state.use_api
        st.rerun()

use_api = st.session_state.use_api
mode_color = GREEN if use_api else YELLOW
mode_text  = "AI mode active — Groq LLaMA 3 will analyse each complaint" if use_api \
             else "Rule-based mode — fast local logic, no API calls needed"
st.markdown(f'<div class="info-card">{"🤖" if use_api else "⚙️"} <b>{"AI Mode"if use_api else "Rule-Based Mode"}:</b> {mode_text}</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def call_groq(messages):
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {GROQ_API_KEY}"}
    payload = {"model": GROQ_MODEL, "messages": messages, "max_tokens": 1024, "temperature": 0.1}
    resp = requests.post(GROQ_URL, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]

def parse_json_safe(text):
    try:
        clean = re.sub(r"```json|```", "", text).strip()
        return json.loads(clean)
    except Exception:
        return None

def is_blank(v):
    return v is None or str(v).strip() == ""

def trim(v, n=200):
    s = str(v).strip()
    return s[:n] + "…" if len(s) > n else s


# ── Rule-based severity ──────────────────────────────────────────────────────
SERIOUS_KW  = ["death","died","passed away","deceased","fatality","fatal","life-threatening",
                "critical","negligence","malpractice","surgery error","wrong medication",
                "overdose","sepsis","cardiac arrest","emergency refused"]
MAJOR_KW    = ["admitted","icu","intensive care","misdiagnosed","misplaced","infection",
                "procedure failed","delayed surgery","wrong treatment","refused treatment",
                "no doctor","unconscious","collapsed"]
MODERATE_KW = ["delay","waiting","long wait","not attended","rude","staff attitude",
                "medicine not available","appointment","billing","overcharged","no wheelchair",
                "not informed","missing file","missing report","missing documents"]
MINOR_KW    = ["minor","small","feedback","suggestion","parking","food","cleanliness",
                "noise","general","inquiry","information"]

def rule_severity(text):
    t = text.lower()
    if any(k in t for k in SERIOUS_KW):  return "Serious"
    if any(k in t for k in MAJOR_KW):    return "Major"
    if any(k in t for k in MODERATE_KW): return "Moderate"
    return "Minor"

def rule_harm(text, category=""):
    t = (text + " " + category).lower()
    harm_kw = ["death","died","fatal","life-threatening","refused treatment","emergency refused",
                "wrong medication","overdose","infection","sepsis","critical","collapse",
                "delay","procedure failed","misdiagnosed","wrong treatment","surgery error"]
    return "Yes" if any(k in t for k in harm_kw) else "No"

def rule_affinity(category, text):
    cat = category.lower()
    t   = text.lower()
    if "wheelchair" in t or "equipment" in t:   return "Equipment Issue"
    if "staff attitude" in cat or "rude" in t:  return "Staff Conduct"
    if "delay" in cat or "wait" in t:            return "Service Delay"
    if "quality" in cat or "care" in t:          return "Care Quality"
    if "communication" in cat:                   return "Communication"
    if "medication" in t or "medicine" in t:     return "Medication Issue"
    if "appointment" in t:                       return "Appointment Issue"
    if "billing" in t or "payment" in t:         return "Billing Issue"
    if "cleanliness" in t or "hygiene" in t:     return "Hygiene Issue"
    if "transport" in cat:                        return "Transport Issue"
    if "missing" in t or "misplaced" in t:       return "Record Issue"
    if "room" in t or "ward" in t:               return "Facility Issue"
    return "General Complaint"


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 – UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="section-head"><div class="step-num">1</div><div class="step-label">Upload Workbook</div></div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Drop your Excel or CSV file here", type=["xlsx","xls","csv"],
                              label_visibility="collapsed")
if uploaded:
    st.session_state.wb_bytes  = uploaded.read()
    st.session_state.file_name = uploaded.name
    st.session_state.result_df = None
    st.session_state.log       = []
    st.markdown(f'<div class="info-card">📂 <b>{uploaded.name}</b> loaded successfully</div>', unsafe_allow_html=True)

if not st.session_state.wb_bytes:
    st.markdown('<div class="warn-card">⬆ Please upload an Excel workbook to begin.</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="omac-footer"><strong>OMAC Complaint Intelligence</strong> · Developed by <strong>S M Baqir</strong></div>', unsafe_allow_html=True)
    st.stop()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 – SHEET
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="section-head"><div class="step-num">2</div><div class="step-label">Select Sheet</div></div>', unsafe_allow_html=True)

fname    = st.session_state.file_name.lower()
wb_bytes = st.session_state.wb_bytes

if fname.endswith(".csv"):
    sheet_name = "Sheet1"
else:
    wb_peek    = load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    sheet_name = st.selectbox("Choose a sheet", wb_peek.sheetnames)
    wb_peek.close()

st.session_state.sheet_name = sheet_name

if fname.endswith(".csv"):
    df = pd.read_csv(io.BytesIO(wb_bytes), dtype=str).fillna("")
else:
    df = pd.read_excel(io.BytesIO(wb_bytes), sheet_name=sheet_name, dtype=str).fillna("")

st.session_state.df = df
st.caption(f"📋  {len(df)} rows · {len(df.columns)} columns")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 – CONFIGURE
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="section-head"><div class="step-num">3</div><div class="step-label">Configure Columns</div></div>', unsafe_allow_html=True)

all_cols = df.columns.tolist()

# ── Complaint description column (what AI reads) ──────────────────────────────
st.markdown(f'<div class="card">', unsafe_allow_html=True)
st.markdown("**📝 Complaint Description Column** — the column the AI reads to understand each complaint")
desc_col_guess = next((c for c in all_cols if any(k in c.lower() for k in
    ["description","complaint","detail","narrat","text","feedback","remark"])), all_cols[0])
desc_col = st.selectbox("Select complaint description column", all_cols,
                         index=all_cols.index(desc_col_guess))
st.markdown("</div>", unsafe_allow_html=True)

# ── Additional context columns ────────────────────────────────────────────────
st.markdown(f'<div class="card">', unsafe_allow_html=True)
st.markdown("**📎 Additional Context Columns** — other columns the AI can use as context (department, category, etc.)")
other_cols   = [c for c in all_cols if c != desc_col]
fill_rate    = {c: (df[c].replace("","").apply(bool).sum() / max(len(df),1)) for c in other_cols}
smart_ctx    = [c for c, r in sorted(fill_rate.items(), key=lambda x:-x[1]) if r > 0.1][:4]
ctx_cols     = st.multiselect("Select context columns (optional)", options=other_cols, default=smart_ctx)
st.markdown("</div>", unsafe_allow_html=True)

# ── Output columns to fill ────────────────────────────────────────────────────
st.markdown(f'<div class="card">', unsafe_allow_html=True)
st.markdown("**🎯 Columns to Fill** — select which outputs the AI should generate")

tc1, tc2, tc3 = st.columns(3)

with tc1:
    sev_col_guess = next((c for c in all_cols if "sever" in c.lower()), None)
    do_severity   = st.checkbox("**Severity**", value=True)
    severity_col  = st.selectbox("Write severity to column", ["— create new column —"] + all_cols,
                                  index=0 if not sev_col_guess else (["— create new column —"]+all_cols).index(sev_col_guess),
                                  disabled=not do_severity)
    if do_severity:
        st.caption("Values: Serious / Major / Moderate / Minor")

with tc2:
    aff_col_guess = next((c for c in all_cols if "affin" in c.lower()), None)
    do_affinity   = st.checkbox("**Affinity**", value=True)
    affinity_col  = st.selectbox("Write affinity to column", ["— create new column —"] + all_cols,
                                  index=0 if not aff_col_guess else (["— create new column —"]+all_cols).index(aff_col_guess),
                                  disabled=not do_affinity)
    if do_affinity:
        st.caption("2–3 word situational category label")

with tc3:
    harm_col_guess = next((c for c in all_cols if "harm" in c.lower()), None)
    do_harm        = st.checkbox("**Harm Tagging**", value=True)
    harm_col       = st.selectbox("Write harm tag to column", ["— create new column —"] + all_cols,
                                   index=0 if not harm_col_guess else (["— create new column —"]+all_cols).index(harm_col_guess),
                                   disabled=not do_harm)
    if do_harm:
        st.caption("Values: Yes / No")

st.markdown("</div>", unsafe_allow_html=True)

if not (do_severity or do_affinity or do_harm):
    st.markdown('<div class="warn-card">⚠ Select at least one output column above.</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 – RUN
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="section-head"><div class="step-num">4</div><div class="step-label">Run Analysis</div></div>', unsafe_allow_html=True)

# Count rows that need work
needs_work = 0
for i in range(len(df)):
    row_needs = False
    if do_severity and severity_col == "— create new column —": row_needs = True
    elif do_severity and is_blank(df.at[i, severity_col]):      row_needs = True
    if do_affinity and affinity_col == "— create new column —": row_needs = True
    elif do_affinity and is_blank(df.at[i, affinity_col]):      row_needs = True
    if do_harm and harm_col == "— create new column —":         row_needs = True
    elif do_harm and is_blank(df.at[i, harm_col]):              row_needs = True
    if row_needs: needs_work += 1

rc1, rc2, rc3 = st.columns([2,1,1])
with rc1:
    run_btn = st.button(
        f"{'🤖 Run AI Analysis' if use_api else '⚙️ Run Rule-Based Analysis'}  ({needs_work} rows)",
        type="primary",
        disabled=(not (do_severity or do_affinity or do_harm) or needs_work == 0),
        use_container_width=True,
    )
with rc2:
    st.metric("Rows to process", needs_work)
with rc3:
    st.metric("Total rows", len(df))


# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════
if run_btn:
    st.session_state.log = []
    result_df = df.copy().reset_index(drop=True)

    # Ensure output columns exist
    sev_out  = severity_col if severity_col != "— create new column —" else "Severity"
    aff_out  = affinity_col if affinity_col != "— create new column —" else "Affinity"
    harm_out = harm_col     if harm_col     != "— create new column —" else "Harm"

    for col in [sev_out, aff_out, harm_out]:
        if col not in result_df.columns:
            result_df[col] = ""

    work_rows = [i for i in range(len(result_df)) if not is_blank(result_df.at[i, desc_col])]

    progress_bar    = st.progress(0, text="Starting…")
    log_placeholder = st.empty()
    logs = []

    def render_log():
        html = '<div class="log-box">'
        for l in logs[-50:]:
            cls  = {"success":"log-success","error":"log-error"}.get(l["t"],"log-info")
            icon = {"success":"✓","error":"✗"}.get(l["t"],"›")
            html += f'<div class="{cls}">{icon} {l["m"]}</div>'
        html += "</div>"
        log_placeholder.markdown(html, unsafe_allow_html=True)

    total = len(work_rows)
    done  = 0

    logs.append({"m": f"Mode: {'AI (Groq)' if use_api else 'Rule-Based'}  |  {total} rows to process", "t": "info"})
    render_log()

    # ── RULE-BASED MODE ──────────────────────────────────────────────────────
    if not use_api:
        for i, ri in enumerate(work_rows):
            desc = str(result_df.at[ri, desc_col])
            cat  = " ".join(str(result_df.at[ri, c]) for c in ctx_cols if c in result_df.columns)

            if do_severity:
                sev = rule_severity(desc + " " + cat)
                result_df.at[ri, sev_out] = sev

            if do_harm:
                hrm = rule_harm(desc, cat)
                result_df.at[ri, harm_out] = hrm

            if do_affinity:
                aff = rule_affinity(cat, desc)
                result_df.at[ri, aff_out] = aff

            sev_v  = result_df.at[ri, sev_out]  if do_severity else "—"
            aff_v  = result_df.at[ri, aff_out]  if do_affinity else "—"
            harm_v = result_df.at[ri, harm_out] if do_harm     else "—"
            logs.append({"m": f"Row {ri+2}: Severity={sev_v}  Affinity={aff_v}  Harm={harm_v}", "t": "success"})

            done += 1
            progress_bar.progress(done/total, text=f"{done}/{total} rows processed ({int(done/total*100)}%)")
            if done % 5 == 0 or done == total:
                render_log()

    # ── AI MODE ──────────────────────────────────────────────────────────────
    else:
        outputs_needed = []
        if do_severity: outputs_needed.append(("severity", sev_out,
            'One word only from: Serious, Major, Moderate, Minor.\n'
            'Serious=death/life-threatening/fatal negligence. '
            'Major=ICU/wrong treatment/refused emergency. '
            'Moderate=delay/rude staff/missing docs/unavailable medicine. '
            'Minor=minor inconvenience/feedback/suggestion.'))
        if do_affinity: outputs_needed.append(("affinity", aff_out,
            '2-3 words MAX describing the situational category of this complaint. '
            'Examples: "Wheelchair Unavailability", "Lab Test Delay", "Staff Misconduct", '
            '"Medication Shortage", "Appointment Issue", "Record Misplacement", '
            '"Care Quality", "Emergency Access". Be specific and concise.'))
        if do_harm:     outputs_needed.append(("harm", harm_out,
            'Yes or No only. Yes if: patient death, life-threatening delay, wrong/missed treatment, '
            'refused emergency, severe medical error, critical medication error. No for: '
            'staff attitude, minor delays, billing, general inconvenience.'))

        system_prompt = (
            "You are a complaint intelligence analyst for a hospital. "
            "Analyse each complaint and return ONLY a JSON array with no explanation or markdown.\n"
            "Fields to return per row:\n" +
            "\n".join(f'- "{k}": {desc}' for k,_,desc in outputs_needed) +
            '\n- "_rowIdx": (keep exactly as given)\n'
            f'Format: [{{"_rowIdx": N, {", ".join(chr(34)+k+chr(34)+": ..." for k,_,__ in outputs_needed)}}}]'
        )

        for b in range(0, total, SAFE_BATCH):
            chunk = work_rows[b: b + SAFE_BATCH]
            rows_payload = []
            for ri in chunk:
                row_data = {"complaint": trim(result_df.at[ri, desc_col], 250)}
                for c in ctx_cols:
                    if c in result_df.columns:
                        row_data[c] = trim(result_df.at[ri, c], 80)
                rows_payload.append({"_rowIdx": ri, "data": row_data})

            user_msg = (
                f"Analyse these {len(chunk)} complaints. Return exactly {len(chunk)} JSON items.\n"
                + json.dumps(rows_payload, ensure_ascii=False)
            )

            batch_num = b // SAFE_BATCH + 1
            logs.append({"m": f"Batch {batch_num}: rows {chunk[0]+2}–{chunk[-1]+2} ({len(chunk)} rows)…", "t": "info"})
            render_log()

            try:
                raw     = call_groq([{"role":"system","content":system_prompt},
                                     {"role":"user",  "content":user_msg}])
                parsed  = parse_json_safe(raw)

                if parsed and isinstance(parsed, list):
                    filled = set()
                    for item in parsed:
                        ri  = int(item.get("_rowIdx", -1))
                        if ri not in chunk: continue
                        for key, col_name, _ in outputs_needed:
                            val = str(item.get(key, "")).strip()
                            if val:
                                result_df.at[ri, col_name] = val
                        filled.add(ri)
                        sev_v  = result_df.at[ri, sev_out]  if do_severity else "—"
                        aff_v  = result_df.at[ri, aff_out]  if do_affinity else "—"
                        harm_v = result_df.at[ri, harm_out] if do_harm     else "—"
                        logs.append({"m": f"Row {ri+2}: Severity={sev_v}  Affinity={aff_v}  Harm={harm_v}", "t": "success"})

                    # Retry skipped rows individually
                    for ri in [r for r in chunk if r not in filled]:
                        logs.append({"m": f"Retrying row {ri+2}…", "t": "info"})
                        render_log()
                        single = {"_rowIdx": ri, "data": {"complaint": trim(result_df.at[ri, desc_col], 250)}}
                        retry_msg = f"Analyse 1 complaint. Return exactly 1 JSON item.\n{json.dumps([single])}"
                        try:
                            r2 = call_groq([{"role":"system","content":system_prompt},
                                            {"role":"user",  "content":retry_msg}])
                            p2 = parse_json_safe(r2)
                            if p2 and isinstance(p2, list) and p2:
                                item = p2[0]
                                for key, col_name, _ in outputs_needed:
                                    val = str(item.get(key,"")).strip()
                                    if val: result_df.at[ri, col_name] = val
                                logs.append({"m": f"Row {ri+2} (retry): ok", "t": "success"})
                            else:
                                logs.append({"m": f"Row {ri+2}: no response after retry", "t": "error"})
                        except Exception as e2:
                            logs.append({"m": f"Row {ri+2} retry error: {str(e2)[:80]}", "t": "error"})
                else:
                    logs.append({"m": f"Batch {batch_num} parse failed — switching to rule-based fallback", "t": "error"})
                    for ri in chunk:
                        desc = str(result_df.at[ri, desc_col])
                        cat  = " ".join(str(result_df.at[ri, c]) for c in ctx_cols if c in result_df.columns)
                        if do_severity: result_df.at[ri, sev_out]  = rule_severity(desc + " " + cat)
                        if do_harm:     result_df.at[ri, harm_out] = rule_harm(desc, cat)
                        if do_affinity: result_df.at[ri, aff_out]  = rule_affinity(cat, desc)
                        logs.append({"m": f"Row {ri+2}: filled via rules (fallback)", "t": "info"})

            except Exception as e:
                logs.append({"m": f"Batch {batch_num} error: {str(e)[:100]} — using rule-based fallback", "t": "error"})
                for ri in chunk:
                    desc = str(result_df.at[ri, desc_col])
                    cat  = " ".join(str(result_df.at[ri, c]) for c in ctx_cols if c in result_df.columns)
                    if do_severity: result_df.at[ri, sev_out]  = rule_severity(desc + " " + cat)
                    if do_harm:     result_df.at[ri, harm_out] = rule_harm(desc, cat)
                    if do_affinity: result_df.at[ri, aff_out]  = rule_affinity(cat, desc)
                    logs.append({"m": f"Row {ri+2}: filled via rules (fallback)", "t": "info"})

            done += len(chunk)
            progress_bar.progress(done/total, text=f"{done}/{total} rows ({int(done/total*100)}%)")
            render_log()

    progress_bar.progress(1.0, text="✓ Complete!")
    logs.append({"m": "All done! Download your workbook below.", "t": "success"})
    render_log()
    st.session_state.result_df = result_df
    st.session_state.log       = logs
    st.session_state["sev_out"]  = sev_out  if do_severity else None
    st.session_state["aff_out"]  = aff_out  if do_affinity else None
    st.session_state["harm_out"] = harm_out if do_harm     else None


# ═══════════════════════════════════════════════════════════════════════════════
# DOWNLOAD + PREVIEW
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.result_df is not None:
    result_df = st.session_state.result_df
    s_out = st.session_state.get("sev_out")
    a_out = st.session_state.get("aff_out")
    h_out = st.session_state.get("harm_out")

    st.markdown('<div class="success-card">✓ Analysis complete! Download your workbook below.</div>', unsafe_allow_html=True)
    st.markdown("")

    # ── Stats summary ──────────────────────────────────────────────────────────
    if s_out and s_out in result_df.columns:
        sc1, sc2, sc3, sc4 = st.columns(4)
        counts = result_df[s_out].value_counts()
        sc1.metric("🔴 Serious",  counts.get("Serious",  0))
        sc2.metric("🟠 Major",    counts.get("Major",    0))
        sc3.metric("🟡 Moderate", counts.get("Moderate", 0))
        sc4.metric("🟢 Minor",    counts.get("Minor",    0))

    if h_out and h_out in result_df.columns:
        harm_yes = (result_df[h_out] == "Yes").sum()
        harm_no  = (result_df[h_out] == "No").sum()
        hc1, hc2 = st.columns(2)
        hc1.metric("⚠️ Harm: Yes", harm_yes)
        hc2.metric("✅ Harm: No",  harm_no)

    st.markdown("")

    # ── Build output file ──────────────────────────────────────────────────────
    out = io.BytesIO()
    orig_df = st.session_state.df.reset_index(drop=True)

    if fname.endswith(".csv"):
        result_df.to_csv(out, index=False)
        out.seek(0)
        dl_name = st.session_state.file_name.replace(".csv","_AI_tagged.csv")
        mime = "text/csv"
    else:
        orig_wb = load_workbook(io.BytesIO(st.session_state.wb_bytes))
        ws      = orig_wb[st.session_state.sheet_name]

        # Extend headers if new columns were added
        existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        col_map = {str(h): i+1 for i, h in enumerate(existing_headers) if h is not None}
        next_col = ws.max_column + 1
        for new_col in result_df.columns:
            if new_col not in col_map:
                ws.cell(1, next_col).value = new_col
                col_map[new_col] = next_col
                next_col += 1

        # Write values
        tagged_cols = [c for c in [s_out, a_out, h_out] if c]
        for ri, row in result_df.iterrows():
            for col_name in tagged_cols:
                if col_name not in result_df.columns: continue
                orig_val = orig_df.at[ri, col_name] if col_name in orig_df.columns else ""
                new_val  = row[col_name]
                if str(new_val).strip() and (is_blank(orig_val) or col_name not in orig_df.columns):
                    excel_row = ri + 2
                    excel_col = col_map.get(col_name)
                    if excel_col:
                        ws.cell(row=excel_row, column=excel_col).value = new_val

        orig_wb.save(out)
        out.seek(0)
        base    = st.session_state.file_name.rsplit(".",1)[0]
        dl_name = f"{base}_AI_tagged.xlsx"
        mime    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    st.download_button(
        label="⬇  Download Tagged Workbook",
        data=out, file_name=dl_name, mime=mime,
        type="primary", use_container_width=True,
    )

    # ── Preview ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Preview — tagged results**")
    preview_cols = [c for c in [s_out, a_out, h_out] if c and c in result_df.columns]
    ctx_preview  = [c for c in (ctx_cols if "ctx_cols" in dir() else []) if c in result_df.columns][:3]
    show_cols    = preview_cols + ctx_preview
    if show_cols:
        st.dataframe(result_df[show_cols].head(20), use_container_width=True, height=380)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown(
    f'<div class="omac-footer"><strong>OMAC Complaint Intelligence</strong> &nbsp;·&nbsp; Developed by <strong>S M Baqir</strong></div>',
    unsafe_allow_html=True,
)
